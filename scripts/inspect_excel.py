"""One-shot Excel inspection. Safe to delete after import is verified."""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from pathlib import Path

XLSX = Path(r'c:\Users\hocch\OneDrive\画像\ドキュメント\南アフリカアプリ\全国動物病院地図アプリ\全国動物病院リスト.xlsx')

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb['all_vets']

rows_iter = ws.iter_rows(values_only=True)
header = next(rows_iter)

print('=== HEADERS ===')
for i, h in enumerate(header):
    print(f'  [{i}] {h}')
print()

print('=== 3 SAMPLE ROWS (focus on URL, business hours, reviews) ===')
samples = []
for i, row in enumerate(rows_iter):
    if i >= 3: break
    samples.append(row)
    print(f'--- Row {i+1} ---')
    for j, cell in enumerate(row):
        val = (str(cell) if cell is not None else '')
        print(f'  [{j}] {header[j]}: {val[:300]}')
    print()

print('=== GOOGLE MAPS URL COLUMN (10 samples) ===')
url_col = 9  # GoogleマップURL
wb2 = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws2 = wb2['all_vets']
rows_iter2 = ws2.iter_rows(values_only=True)
next(rows_iter2)  # skip header
url_samples = []
for i, row in enumerate(rows_iter2):
    if i >= 20: break
    url_samples.append(row[url_col])
for i, u in enumerate(url_samples[:10]):
    print(f'  [{i}] {u}')
print()

print('=== URL FORMAT STATS (first 500 rows) ===')
import re
at_coords_re = re.compile(r'@(-?\d+\.\d+),(-?\d+\.\d+)')
cid_re = re.compile(r'[?&]cid=')
pid_re = re.compile(r'place_id[:=]')
count_at = count_cid = count_pid = count_other = total = 0
wb3 = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws3 = wb3['all_vets']
rows_iter3 = ws3.iter_rows(values_only=True)
next(rows_iter3)
for i, row in enumerate(rows_iter3):
    if i >= 500: break
    total += 1
    url = str(row[url_col] or '')
    if at_coords_re.search(url):
        count_at += 1
    elif cid_re.search(url):
        count_cid += 1
    elif pid_re.search(url):
        count_pid += 1
    else:
        count_other += 1
print(f'  Total checked: {total}')
print(f'  @lat,lng style (extractable): {count_at}')
print(f'  cid= style: {count_cid}')
print(f'  place_id style: {count_pid}')
print(f'  Other: {count_other}')

print()
print('=== REVIEW COLUMN (1 full sample) ===')
review_col = 8  # 口コミ(最大5件)
if samples:
    review_text = str(samples[0][review_col] or '')
    print(review_text[:2000])
    print('---')
    print(f'Length: {len(review_text)} chars')
    # Detect delimiters
    print(f'Newlines: {review_text.count(chr(10))}')
    print(f'Pipes |: {review_text.count("|")}')
    print(f'Triple dashes ---: {review_text.count("---")}')
    print(f'Star (★): {review_text.count(chr(0x2605))}')

print()
print('=== BUSINESS HOURS COLUMN (1 full sample) ===')
hours_col = 7
if samples:
    h_text = str(samples[0][hours_col] or '')
    print(h_text)
    print('---')
    print(f'Length: {len(h_text)} chars, newlines: {h_text.count(chr(10))}')

print()
print('=== PREFECTURE / AREA UNIQUE COUNT ===')
pref_col = 11
area_col = 10
prefs = set()
areas = set()
wb4 = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws4 = wb4['all_vets']
ri = ws4.iter_rows(values_only=True)
next(ri)
total_rows = 0
missing_url = missing_pref = missing_phone = missing_web = missing_hours = 0
rating_min, rating_max = 999, -1
review_min, review_max = 999999, 0
for row in ri:
    total_rows += 1
    if row[pref_col]: prefs.add(row[pref_col])
    if row[area_col]: areas.add(row[area_col])
    if not row[url_col]: missing_url += 1
    if not row[pref_col]: missing_pref += 1
    if not row[5]: missing_phone += 1
    if not row[6]: missing_web += 1
    if not row[7]: missing_hours += 1
    try:
        r = float(row[3]) if row[3] is not None else None
        if r is not None:
            rating_min = min(rating_min, r)
            rating_max = max(rating_max, r)
    except: pass
    try:
        rc = int(row[4]) if row[4] is not None else None
        if rc is not None:
            review_min = min(review_min, rc)
            review_max = max(review_max, rc)
    except: pass
print(f'Total rows: {total_rows}')
print(f'Prefectures (unique): {len(prefs)} -> {sorted(prefs)[:5]}... total {len(prefs)}')
print(f'Areas (unique): {len(areas)}')
print(f'Missing: url={missing_url}, pref={missing_pref}, phone={missing_phone}, web={missing_web}, hours={missing_hours}')
print(f'Rating range: {rating_min} - {rating_max}')
print(f'Review count range: {review_min} - {review_max}')
